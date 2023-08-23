import openpyxl
#Format of the Excel files:
# File---->Workbook---->Sheets--->Rows---->Cells.

#Step 1: first give the path of the excel file
file = "D:/selenium/userfile.xlsx"

#Step2: We need to extract the Workbook from the excel
workbook = openpyxl.load_workbook(file)

#Step3: extract the sheet out of all the sheets.
sheet = workbook["Sheet1"]

#Step4: get all the rows and columns(Read the data)
#we need to use looping system to read the rows and columns
#how to find the number of rows and columns in the sheet
#(sheet.max_row)      :::::: sheet.max_column

rows = sheet.max_row
cols = sheet.max_column

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value, end ='         ')
    print()


