import openpyxl

#this is for the same data in the excel file

# file = "D:/selenium/data.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet1"]
# #sheet = workbook.active
# #if we have only one active sheet in the excel file.
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value = "welcome"
# workbook.save(file)


# this is the file for different data.


file = "D:/selenium/data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet2"]
#sheet = workbook.active
#if we have only one active sheet in the excel file.

sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "smith"
sheet.cell(1,3).value = "engineer"

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = "john"
sheet.cell(2,3).value = "developer"

sheet.cell(3,1).value = 567
sheet.cell(3,2).value = "david"
sheet.cell(3,3).value = "manager"

workbook.save(file)
# #save the file.
#
# data =[('Emp Id', 'Emp Name', 'Designation'),
#        (1, 'XYZ', 'Manager'),
#        (2, 'ABC', 'Consultant')]
# for i in data:
#     sheet3.append(i)
#
# workbook.save(file)

